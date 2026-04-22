#!/usr/bin/env python3
"""
Nyingmapa Calendar — Excel to JSON converter
Input:  source/Nyingmapa_Calendar.xlsx  (auto-detected from source/ folder)
Output: assets/data/ directory with monthly JSON + reference files

Usage:
  python3 scripts/excel_to_json.py
  python3 scripts/excel_to_json.py --input path/to/file.xlsx
  python3 scripts/excel_to_json.py --output path/to/data/
"""

import json
import sys
import argparse
from pathlib import Path
from datetime import datetime
import openpyxl


# ─────────────────────────────────────────────
# Config
# ─────────────────────────────────────────────

MONTH_ORDER = ["JAN","FEB","MAR","APR","MAY","JUN",
               "JUL","AUG","SEP","OCT","NOV","DEC"]


# Maps column prefix in Excel → key in JSON astrology object
ASTROLOGY_FIELDS = [
    "naga_day",
    "flag_day",
    "fire_ritual",
    "torma_offering",
    "empty_vase",
    "hair_cutting",
    "inauspicious_day",
    "daily_restriction",
    "auspicious_time",
]

# Maps Excel col prefix → astrology key
ASTROLOGY_COL_MAP = {
    "naga_day":          "naga_day",
    "flag_day":          "flag_day",
    "fire_ritual":       "fire_ritual",
    "torma_offering":    "torma_offering",
    "empty_vase":        "empty_vase",
    "hair_cutting":      "hair_cutting",
    "inauspicious_day":  "inauspicious_day",
    "daily_restriction": "daily_restriction",
    "auspicious_time":   "auspicious_time",
}

EVENT_PREFIXES = ["first_event", "second_event", "third_event", "fourth_event"]


# ─────────────────────────────────────────────
# Utility
# ─────────────────────────────────────────────

def clean(val):
    """Convert cell value to clean Python value. null/None → None."""
    if val is None:
        return None
    if isinstance(val, str):
        stripped = val.strip()
        if stripped.lower() in ("null", "none", ""):
            return None
        return stripped
    if isinstance(val, float):
        if val == int(val):
            return int(val)
        return val
    return val


def col_map(headers):
    """Build {header_name: index} dict from header list."""
    return {h: i for i, h in enumerate(headers) if h}


def get(row, cm, *keys):
    """Get first matching key from column map."""
    for k in keys:
        if k in cm and cm[k] < len(row):
            v = clean(row[cm[k]])
            if v is not None:
                return v
    return None


def normalize_image(val):
    """
    Normalize image key:
    - int/float 1-8 → int (daily rotation, Flutter renders as "1.webp")
    - string → lowercase, strip spaces, remove extension
    """
    if val is None:
        return None
    if isinstance(val, (int, float)):
        n = int(val)
        return n if 1 <= n <= 8 else None
    s = str(val).strip()
    if not s or s.lower() in ("null", "none"):
        return None
    # Remove file extension
    for ext in (".png", ".jpg", ".jpeg", ".webp", ".PNG", ".JPG"):
        if s.endswith(ext):
            s = s[: -len(ext)]
    return s.lower().strip()


def make_date_str(year, month, day):
    """Build ISO date string from year(int), month(str 'FEB'), day(int)."""
    try:
        m = MONTH_ORDER.index(str(month).upper().strip()) + 1
        return f"{int(year):04d}-{m:02d}-{int(day):02d}"
    except Exception:
        return None


def tibetan_date_display(day, month, year):
    """Build display string like 'Day 10th, Month 1st, 2153'."""
    def ordinal(n):
        n = int(n)
        s = {1: "st", 2: "nd", 3: "rd"}.get(n % 10 if n % 100 not in (11,12,13) else 0, "th")
        return f"{n}{s}"
    try:
        return f"Day {ordinal(day)}, Month {ordinal(month)}, {int(year)}"
    except Exception:
        return None


def gregorian_display(month, day, year):
    """Build display string like 'Feb 18, 2026'."""
    try:
        m = MONTH_ORDER.index(str(month).upper().strip()) + 1
        dt = datetime(int(year), m, int(day))
        return dt.strftime("%b %-d, %Y")
    except Exception:
        return None


# ─────────────────────────────────────────────
# Excel Reader
# ─────────────────────────────────────────────

def find_header_row(ws, min_cols=5):
    """Find the row index (0-based) that looks like the header row."""
    for i, row in enumerate(ws.iter_rows(max_row=10, values_only=True)):
        if sum(1 for c in row if c is not None) >= min_cols:
            return i
    return 0


def load_sheet_with_headers(ws):
    """Return (headers_list, data_rows_list) for a sheet."""
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, row in enumerate(rows[:10]):
        count = sum(1 for c in row if c is not None)
        if count >= 5:
            header_idx = i
            break
    if header_idx is None:
        header_idx = 0

    headers = list(rows[header_idx])
    data = rows[header_idx + 1:]
    return headers, data


# ─────────────────────────────────────────────
# Daily Parser
# ─────────────────────────────────────────────

def parse_daily_row(row, cm):
    """Parse one data row from daily_full → day dict."""

    # ── Gregorian ──
    g_year  = clean(row[cm["gregorian_year_en"]])
    g_month = clean(row[cm["gregorian_month_en"]])
    g_day   = clean(row[cm["gregorian_date_en"]])
    g_dow   = clean(row[cm["gregorian_day_en"]])

    if not all([g_year, g_month, g_day]):
        return None  # skip incomplete rows

    date_str = make_date_str(g_year, g_month, g_day)
    if not date_str:
        return None

    gregorian = {
        "year":         int(g_year),
        "month":        str(g_month).upper(),
        "day":          int(g_day),
        "day_of_week":  g_dow,
        "year_bo":      get(row, cm, "gregorian_year_bo"),
        "month_bo":     get(row, cm, "gregorian_month_bo"),
        "date_bo":      get(row, cm, "gregorian_date_bo"),
        "day_of_week_bo": get(row, cm, "gregorian_day_bo"),
    }

    # ── Tibetan ──
    t_year  = clean(row[cm.get("tibetan_year_en", -1)]) if "tibetan_year_en" in cm else None
    t_month = clean(row[cm.get("tibetan_month_en", -1)]) if "tibetan_month_en" in cm else None
    t_day   = clean(row[cm.get("tibetan_day_en", -1)]) if "tibetan_day_en" in cm else None

    tibetan = {
        "year":          int(t_year) if t_year else None,
        "month":         int(t_month) if t_month else None,
        "month_name":    get(row, cm, "tibetan_month_name_en"),
        "animal_month":  get(row, cm, "animal_month_en"),
        "day":           int(t_day) if t_day else None,
        "year_bo":       get(row, cm, "tibetan_year_bo"),
        "month_bo":      get(row, cm, "tibetan_month_bo"),
        "month_name_bo": get(row, cm, "tibetan_month_name_bo"),
        "animal_month_bo": get(row, cm, "animal_month_bo"),
        "day_bo":        get(row, cm, "tibetan_day_bo"),
        # Year name from Excel (e.g. "Fire Horse" / "མེ་ཕོ་རྟ་ལོ།")
        "year_name_en":  get(row, cm, "tibetan_year_name_en"),
        "year_name_bo":  get(row, cm, "tibetan_year_name_bo"),
    }

    # ── Auspicious day ──
    ausp_name = get(row, cm, "auspicious_day_name_en")
    # handle possible space in col name
    ausp_name_bo = None
    for k in ("auspicious_day_name_bo", "auspicious_day _name_bo"):
        if k in cm:
            ausp_name_bo = clean(row[cm[k]])
            break

    auspicious_day = None
    if ausp_name:
        auspicious_day = {
            "name_en":      ausp_name,
            "name_bo":      ausp_name_bo,
            "short_desc_en": get(row, cm, "auspicious_day_shortdescription_en"),
            "short_desc_bo": get(row, cm, "auspicious_day_shortdescription_bo"),
        }

    # ── Lunar status ──
    lunar_en = get(row, cm, "lunar_status_en")
    lunar_bo = get(row, cm, "lunar_status_bo")
    lunar_status = {"en": lunar_en, "bo": lunar_bo} if lunar_en else None

    # ── Day significance ──
    sig_en = get(row, cm, "day_significane_en")
    sig_bo = get(row, cm, "day_significane_bo")
    day_significance = {"en": sig_en, "bo": sig_bo} if sig_en else None

    # ── Element combo ──
    elem_en = get(row, cm, "element_combo_en")
    elem_bo = get(row, cm, "element_combo_bo")
    elem_mean_en = get(row, cm, "meaning_of_coincidence_en")
    elem_mean_bo = get(row, cm, "meaning_of_coincidence_bo")
    element_combo = None
    if elem_en:
        element_combo = {
            "combo_en":   elem_en,
            "combo_bo":   elem_bo,
            "meaning_en": elem_mean_en,
            "meaning_bo": elem_mean_bo,
        }

    # ── Daily image ──
    daily_img = normalize_image(row[cm["daily_image"]] if "daily_image" in cm else None)

    # ── Astrology cards ──
    astrology = {}
    for prefix, key in ASTROLOGY_COL_MAP.items():
        status_en_col  = f"{prefix}_status_en"
        status_bo_col  = f"{prefix}_status_bo"
        image_col      = f"{prefix}_image"
        popup_col      = f"{prefix}_popup_screen"

        status_en = get(row, cm, status_en_col)
        status_bo = get(row, cm, status_bo_col)
        image     = normalize_image(get(row, cm, image_col))
        popup_ref = get(row, cm, popup_col)

        astrology[key] = {
            "status_en": status_en,
            "status_bo": status_bo,
            "image":     image,
            "popup_ref": popup_ref,
        }

    # ── Events ──
    events = []
    for prefix in EVENT_PREFIXES:
        name_en  = get(row, cm, f"{prefix}_name_en")
        if not name_en:
            continue
        events.append({
            "name_en":    name_en,
            "name_bo":    get(row, cm, f"{prefix}_name_bo"),
            "image":      normalize_image(get(row, cm, f"{prefix}_name_image")),
            "category_en": get(row, cm, f"{prefix}_category_en"),
            "category_bo": get(row, cm, f"{prefix}_category_bo"),
            "details_en":  get(row, cm, f"{prefix}_details_en"),
            "details_bo":  get(row, cm, f"{prefix}_details_bo"),
        })

    return {
        "date": date_str,
        "gregorian": gregorian,
        "tibetan": tibetan,
        "daily_image": daily_img,
        "auspicious_day": auspicious_day,
        "lunar_status": lunar_status,
        "day_significance": day_significance,
        "element_combo": element_combo,
        "astrology": astrology,
        "events": events,
    }


# ─────────────────────────────────────────────
# Index Builders
# ─────────────────────────────────────────────

def build_auspicious_index(days):
    """Build flat list of auspicious days from parsed day objects."""
    result = []
    for day in days:
        if not day.get("auspicious_day"):
            continue
        g  = day["gregorian"]
        t  = day["tibetan"]
        ad = day["auspicious_day"]

        # Derive type_key from name
        type_key = (ad["name_en"] or "")
        type_key = type_key.lower().replace(" ", "_").replace("(", "").replace(")", "").replace("-", "_")
        type_key = "".join(c for c in type_key if c.isalnum() or c == "_").strip("_")

        result.append({
            "date":             day["date"],
            "gregorian_display": gregorian_display(g["month"], g["day"], g["year"]),
            "tibetan_display":  tibetan_date_display(t["day"], t["month"], t["year"]) if t["day"] else None,
            "type_key":         type_key,
            "name_en":          ad["name_en"],
            "name_bo":          ad["name_bo"],
            "short_desc_en":    ad["short_desc_en"],
            "short_desc_bo":    ad["short_desc_bo"],
            "image":            day["daily_image"],
        })
    return result


def build_events_index(days):
    """Build flat list of all events with date info."""
    result = []
    event_id = 0
    for day in days:
        if not day.get("events"):
            continue
        g = day["gregorian"]
        t = day["tibetan"]
        for ev in day["events"]:
            result.append({
                "id":               f"{day['date']}-{event_id}",
                "date":             day["date"],
                "gregorian_display": gregorian_display(g["month"], g["day"], g["year"]),
                "tibetan_display":  tibetan_date_display(t["day"], t["month"], t["year"]) if t["day"] else None,
                "name_en":          ev["name_en"],
                "name_bo":          ev["name_bo"],
                "image":            ev["image"],
                "category_en":      ev["category_en"],
                "category_bo":      ev["category_bo"],
                "details_en":       ev["details_en"],
                "details_bo":       ev["details_bo"],
            })
            event_id += 1
    return result


# ─────────────────────────────────────────────
# Reference Parser
# ─────────────────────────────────────────────

def parse_reference_sheet(ws, popup_ref_key=None):
    """
    Generic parser for reference sheets.
    Returns dict with title + rows.
    """
    rows = list(ws.iter_rows(values_only=True))

    # Find first non-empty row as title, then header row
    title = None
    header_idx = None
    for i, row in enumerate(rows[:10]):
        non_empty = [c for c in row if c is not None]
        if not non_empty:
            continue
        if title is None:
            title = str(non_empty[0]).strip()
        elif header_idx is None and len(non_empty) >= 2:
            header_idx = i
            break

    if header_idx is None:
        header_idx = min(2, len(rows) - 1)

    raw_headers = rows[header_idx]
    headers = []
    for i, h in enumerate(raw_headers):
        if h is not None:
            cleaned = str(h).strip().lower().replace(" ", "_").replace("/", "_").replace("(", "").replace(")", "")
            headers.append((i, cleaned))

    table_rows = []
    for row in rows[header_idx + 1:]:
        if not any(c is not None for c in row):
            continue
        obj = {}
        for col_i, col_name in headers:
            if col_i < len(row):
                v = clean(row[col_i])
                if v is not None:
                    obj[col_name] = v
        if obj:
            table_rows.append(obj)

    result = {
        "title": title,
        "rows":  table_rows,
    }
    if popup_ref_key:
        result["popup_ref"] = popup_ref_key

    return result


POPUP_REF_MAP = {
    "naga_days":             ["naga-sleep_popup_ref", "naga-minor_popup_ref", "naga-major_popup_ref"],
    "flag_days":             "flag_popup_ref",
    "hair_cutting":          "hair_cutting_popup_ref",
    "horse_death":           "horse_death_popup_ref",
    "fire_rituals":          "fire_popup_ref",
    "daily_restrictions":    "daily_restrictions_popup_ref",
    "empty_vase":            "empty_vase_popup_ref",
    "torma_offerings":       "torma_popup_ref",
    "auspicious_timing":     "auspicious_time_popup_ref",
    "life_force_male":       None,
    "life_force_female":     None,
    "eye_twitching":         None,
    "fatal_weekdays":        None,
    "gu_mig":                None,
}


# ─────────────────────────────────────────────
# Writer
# ─────────────────────────────────────────────

class SafeEncoder(json.JSONEncoder):
    """Handle datetime and other non-serializable types."""
    def default(self, obj):
        if isinstance(obj, datetime):
            return obj.isoformat()
        if hasattr(obj, "isoformat"):
            return obj.isoformat()
        try:
            return str(obj)
        except Exception:
            return None


def write_json(path, data):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"), cls=SafeEncoder)
    size_kb = path.stat().st_size / 1024
    print(f"  ✅ {path.name} ({size_kb:.1f} KB)")


# ─────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────
# Flatten nested day → flat format (Flutter CalendarMonthModel schema)
# ─────────────────────────────────────────────
# Build per-day detail JSON (exact format DayDetailModel.fromJson expects)
# ─────────────────────────────────────────────

# Astrology keys that become cards in the detail view (exclude direction/desc fields).
_CARD_KEYS = ["naga_day", "flag_day", "fire_ritual", "hair_cutting", "inauspicious_day"]

# Astrology keys that are direction-based (torma_offering, empty_vase).
_DIRECTION_KEYS = {"torma_offering", "empty_vase"}

# Astrology keys that are description-based (daily_restriction, auspicious_time).
_DESC_KEYS = {"daily_restriction", "auspicious_time"}


def build_day_detail(day: dict) -> dict:
    """
    Build a day-detail JSON object in the exact schema DayDetailModel.fromJson expects.

    Schema:
    {
      "date_key":             "2026-04-03",
      "tibetan_year_name_en": "Fire Horse",
      "image_key":            "losar" | null,

      "gregorian": { year(int), year_bo, month_en(3-letter), month_bo, date(int),
                     date_bo, weekday_en, weekday_bo },
      "tibetan":   { year_en, year_bo, month_en, month_bo, month_name_en, month_name_bo,
                     animal_month_en, animal_month_bo, day_en, day_bo,
                     lunar_status_en, lunar_status_bo },
      "significance": { day_significance_en, day_significance_bo,
                        element_combo_en, element_combo_bo,
                        meaning_of_coincidence_en, meaning_of_coincidence_bo },

      "auspicious_day":      null | { name_en, name_bo, short_description_en, short_description_bo },
      "astrology_cards":     [ { type, status_en, status_bo, image_key, popup_ref } ],
      "torma_offering":      null | { direction_en, image_key, popup_ref },
      "empty_vase":          null | { direction_en, image_key, popup_ref },
      "daily_restriction":   null | { description_en, image_key, popup_ref },
      "auspicious_times":    null | { description_en, image_key, popup_ref },
      "events":              [ { name_en, name_bo, category_en, category_bo,
                                 details_en, details_bo, image_key } ],
    }
    """
    g   = day.get("gregorian", {})
    t   = day.get("tibetan",   {})
    ad  = day.get("auspicious_day")
    ec  = day.get("element_combo")
    ls  = day.get("lunar_status")
    ds  = day.get("day_significance")
    ast = day.get("astrology", {})

    g_month = str(g.get("month", "JAN")).upper().strip()  # already 3-letter abbr from MONTH_ORDER

    # ── Astrology cards ───────────────────────────────────────────────────────
    cards = []
    for key in _CARD_KEYS:
        a = ast.get(key)
        if not isinstance(a, dict):
            continue
        status_en = a.get("status_en")
        if not status_en or str(status_en).lower() in ("not_applicable", "none", "null", ""):
            continue
        cards.append({
            "type":      key,
            "status_en": status_en,
            "status_bo": a.get("status_bo") or "",
            "image_key": a.get("image"),
            "popup_ref": a.get("popup_ref") or key,
        })

    # ── Direction fields (torma_offering, empty_vase) ─────────────────────────
    def _direction_obj(key):
        a = ast.get(key)
        if not isinstance(a, dict):
            return None
        val = a.get("status_en")
        if not val or str(val).lower() in ("not_applicable", "none", "null", ""):
            return None
        return {
            "direction_en": val,
            "image_key":    a.get("image"),
            "popup_ref":    a.get("popup_ref") or key,
        }

    # ── Description fields (daily_restriction, auspicious_time) ──────────────
    def _desc_obj(key, out_key="description_en"):
        a = ast.get(key)
        if not isinstance(a, dict):
            return None
        val = a.get("status_en")
        if not val or str(val).lower() in ("not_applicable", "none", "null", ""):
            return None
        return {
            out_key:     val,
            "image_key": a.get("image"),
            "popup_ref": a.get("popup_ref") or key,
        }

    # ── Events ────────────────────────────────────────────────────────────────
    events = []
    for ev in day.get("events", []):
        events.append({
            "name_en":     ev.get("name_en") or "",
            "name_bo":     ev.get("name_bo") or "",
            "category_en": ev.get("category_en"),
            "category_bo": ev.get("category_bo"),
            "details_en":  ev.get("details_en"),
            "details_bo":  ev.get("details_bo"),
            "image_key":   ev.get("image"),
        })

    return {
        "date_key":             day["date"],
        "tibetan_year_name_en": t.get("year_name_en"),
        "image_key":            str(day["daily_image"]) if day.get("daily_image") is not None else None,

        "gregorian": {
            "year":       g.get("year"),         # int
            "year_bo":    g.get("year_bo"),
            "month_en":   g_month,               # "APR", "DEC", etc.
            "month_bo":   g.get("month_bo"),
            "date":       g.get("day"),           # int
            "date_bo":    g.get("date_bo"),
            "weekday_en": g.get("day_of_week"),
            "weekday_bo": g.get("day_of_week_bo"),
        },

        "tibetan": {
            "year_en":         str(t["year"])   if t.get("year")   is not None else None,
            "year_bo":         t.get("year_bo"),
            "year_name_en":    t.get("year_name_en"),
            "year_name_bo":    t.get("year_name_bo"),
            "month_en":        str(t["month"])  if t.get("month")  is not None else None,
            "month_bo":        t.get("month_bo"),
            "month_name_en":   t.get("month_name"),
            "month_name_bo":   t.get("month_name_bo"),
            "animal_month_en": t.get("animal_month"),
            "animal_month_bo": t.get("animal_month_bo"),
            "day_en":          str(t["day"])    if t.get("day")    is not None else None,
            "day_bo":          t.get("day_bo"),
            "lunar_status_en": ls["en"] if ls else None,
            "lunar_status_bo": ls["bo"] if ls else None,
        },

        "significance": {
            "day_significance_en":       ds["en"] if ds else None,
            "day_significance_bo":       ds["bo"] if ds else None,
            "element_combo_en":          ec["combo_en"]   if ec else None,
            "element_combo_bo":          ec["combo_bo"]   if ec else None,
            "meaning_of_coincidence_en": ec["meaning_en"] if ec else None,
            "meaning_of_coincidence_bo": ec["meaning_bo"] if ec else None,
        },

        "auspicious_day": {
            "name_en":              ad["name_en"],
            "name_bo":              ad.get("name_bo") or "",
            "short_description_en": ad.get("short_desc_en") or "",
            "short_description_bo": ad.get("short_desc_bo") or "",
        } if ad else None,

        "astrology_cards":   cards,
        "torma_offering":    _direction_obj("torma_offering"),
        "empty_vase":        _direction_obj("empty_vase"),
        "daily_restriction": _desc_obj("daily_restriction"),
        # note: Excel key is "auspicious_time", JSON key Flutter reads is "auspicious_times"
        "auspicious_times":  _desc_obj("auspicious_time"),

        "events": events,
    }


# ─────────────────────────────────────────────

def _flatten_day(day: dict) -> dict:
    """Convert nested parse_daily_row output → flat dict that Flutter reads."""
    g  = day.get("gregorian", {})
    t  = day.get("tibetan",   {})
    ad = day.get("auspicious_day")
    ec = day.get("element_combo")
    ls = day.get("lunar_status")
    ast= day.get("astrology", {})

    # astrology_status: keep only status_en value per key (flat map)
    astrology_status = {}
    for k, v in ast.items():
        s = v.get("status_en") if isinstance(v, dict) else None
        if s:
            astrology_status[k] = s

    date_key = day.get("date", "")

    return {
        "date_key":               date_key,
        "gregorian_date":         str(g["day"]) if g.get("day") is not None else None,
        "gregorian_date_bo":      g.get("date_bo"),
        "weekday_en":             g.get("day_of_week"),
        "weekday_bo":             g.get("day_of_week_bo"),
        "tibetan_day_en":         str(t["day"])    if t.get("day")    is not None else None,
        "tibetan_day_bo":         t.get("day_bo"),
        "tibetan_month_en":       str(t["month"])  if t.get("month")  is not None else None,
        "tibetan_month_bo":       t.get("month_bo"),
        "tibetan_month_name_en":  t.get("month_name"),
        "tibetan_month_name_bo":  t.get("month_name_bo"),
        "tibetan_year_en":        str(t["year"])   if t.get("year")   is not None else None,
        "tibetan_year_bo":        t.get("year_bo"),
        "tibetan_year_name_en":   t.get("year_name_en"),
        "tibetan_year_name_bo":   t.get("year_name_bo"),
        "animal_month_en":        t.get("animal_month"),
        "animal_month_bo":        t.get("animal_month_bo"),
        "lunar_status_en":        ls.get("en") if ls else None,
        "image_key":              str(day["daily_image"]) if day.get("daily_image") is not None else None,
        "auspicious_day_name_en": ad["name_en"] if ad else None,
        "auspicious_day_name_bo": ad["name_bo"] if ad else None,
        # ── element combo (DAY column in hero amber panel) ──────────
        "element_combo_en":       ec["combo_en"] if ec else None,
        "element_combo_bo":       ec["combo_bo"] if ec else None,
        # ────────────────────────────────────────────────────────────
        "astrology_status":       astrology_status if astrology_status else None,
        "has_event":              len(day.get("events", [])) > 0,
    }


# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Convert Nyingmapa Calendar Excel to JSON")
    parser.add_argument("--input",  default=None, help="Path to .xlsx file")
    parser.add_argument("--output", default=None, help="Output directory (default: data/ next to script)")
    args = parser.parse_args()

    # ── Find input file ──
    script_dir = Path(__file__).parent
    if args.input:
        xlsx_path = Path(args.input)
    else:
        source_dir = script_dir / ".." / "source"
        candidates = sorted(source_dir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
        if not candidates:
            print("❌ No .xlsx file found in source/. Use --input to specify.")
            sys.exit(1)
        xlsx_path = candidates[0]
        print(f"📂 Using: {xlsx_path.name}")

    output_dir = Path(args.output) if args.output else script_dir / ".." / "assets" / "data"
    output_dir.mkdir(parents=True, exist_ok=True)

    # ── Load workbook ──
    print("\n⏳ Loading workbook...")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # ── Parse daily_full ──
    print("\n📅 Parsing daily_full...")
    ws = wb["daily_full"]
    headers, data_rows = load_sheet_with_headers(ws)
    cm = col_map(headers)

    # Also handle header with space 'auspicious_day _name_bo'
    for i, h in enumerate(headers):
        if h and " " in str(h):
            clean_key = str(h).replace(" ", "_")
            if clean_key not in cm:
                cm[clean_key] = i

    all_days = []
    skipped = 0
    for row in data_rows:
        if not any(v is not None for v in row):
            continue
        parsed = parse_daily_row(row, cm)
        if parsed:
            all_days.append(parsed)
        else:
            skipped += 1

    print(f"  Parsed: {len(all_days)} days, skipped: {skipped}")

    # ── Group by month and write ──
    print("\n📁 Writing monthly files...")
    monthly = {}
    for day in all_days:
        g = day["gregorian"]
        key = (g["year"], g["month"])
        monthly.setdefault(key, []).append(day)

    calendar_dir = output_dir / "calendar"
    for (year, month), days in sorted(monthly.items(), key=lambda x: (x[0][0], MONTH_ORDER.index(x[0][1]))):
        m_num  = MONTH_ORDER.index(month) + 1
        first  = days[0]
        # Year name comes directly from Excel data (tibetan_year_name_en/bo columns)
        year_name_en = first["tibetan"].get("year_name_en") or ''
        year_name_bo = first["tibetan"].get("year_name_bo") or ''
        # ── Flat format (matches Flutter CalendarMonthModel.fromJson) ──
        flat_days = [_flatten_day(d) for d in days]
        write_json(
            calendar_dir / f"{year}_{m_num:02d}.json",
            {
                "year":                year,
                "month":               m_num,
                "gregorian_month_en":  month,
                "gregorian_month_bo":  first["gregorian"].get("month_bo"),
                "year_name_en":        year_name_en,
                "year_name_bo":        year_name_bo,
                "days":                flat_days,
            }
        )

    # ── Per-day detail files ──
    print("\n📄 Writing per-day detail files...")
    detail_dir = output_dir / "day_details"
    detail_dir.mkdir(parents=True, exist_ok=True)
    for day in all_days:
        detail = build_day_detail(day)
        path = detail_dir / f"{day['date']}.json"
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(detail, f, ensure_ascii=False, separators=(",", ":"), cls=SafeEncoder)
    print(f"  ✅ {len(all_days)} day detail files → {detail_dir}/")

    # ── Auspicious index ──
    print("\n⭐ Building auspicious_index.json...")
    ausp_list = build_auspicious_index(all_days)
    print(f"  {len(ausp_list)} auspicious days")
    write_json(
        output_dir / "auspicious_index.json",
        {
            "year": all_days[0]["gregorian"]["year"] if all_days else None,
            "tibetan_year": all_days[0]["tibetan"]["year"] if all_days else None,
            "count": len(ausp_list),
            "days": ausp_list,
        }
    )

    # ── Events index ──
    print("\n🗓️  Building events_index.json...")
    ev_list = build_events_index(all_days)
    # Collect categories
    categories = sorted(set(e["category_en"] for e in ev_list if e.get("category_en")))
    print(f"  {len(ev_list)} events, categories: {categories}")
    write_json(
        output_dir / "events_index.json",
        {
            "year": all_days[0]["gregorian"]["year"] if all_days else None,
            "count": len(ev_list),
            "categories": categories,
            "events": ev_list,
        }
    )

    # ── Reference sheets ──
    print("\n📚 Parsing reference sheets...")
    ref_dir = output_dir / "reference"

    # astrology_cards_ref (astrology_index sheet)
    if "astrology_index" in wb.sheetnames:
        ws_ai = wb["astrology_index"]
        rows_ai = list(ws_ai.iter_rows(values_only=True))
        # Header at row index 2 (0-based)
        hdr_row = next((i for i, r in enumerate(rows_ai) if r and r[0] and str(r[0]).strip().lower() == "system_name_en"), 2)
        hdrs = [str(c).strip().lower().replace(" ", "_") if c else f"col_{i}" for i, c in enumerate(rows_ai[hdr_row])]
        cards = []
        for row in rows_ai[hdr_row + 1:]:
            if not any(v is not None for v in row):
                continue
            obj = {}
            for i, h in enumerate(hdrs):
                if i < len(row) and row[i] is not None:
                    v = clean(row[i])
                    if v is not None:
                        obj[h] = v
            if obj.get("system_name_en"):
                cards.append(obj)
        write_json(ref_dir / "astrology_cards_ref.json", {"cards": cards})

    # auspicious_days_ref
    if "auspicious_days_ref" in wb.sheetnames:
        ws_adr = wb["auspicious_days_ref"]
        rows_adr = list(ws_adr.iter_rows(values_only=True))
        hdrs_adr = [str(c).strip().lower() if c else None for c in rows_adr[0]]
        types = []
        for row in rows_adr[1:]:
            if not any(v is not None for v in row):
                continue
            obj = {}
            for i, h in enumerate(hdrs_adr):
                if h and i < len(row) and row[i] is not None:
                    v = clean(row[i])
                    if v is not None:
                        obj[h] = v
            if obj.get("auspicious_day_name_en"):
                types.append(obj)
        write_json(ref_dir / "auspicious_days_ref.json", {"types": types})

    # daily_astrological_cards_ref
    if "daily_astrological_cards_ref" in wb.sheetnames:
        ws_dacr = wb["daily_astrological_cards_ref"]
        hdrs_dacr, rows_dacr = load_sheet_with_headers(ws_dacr)
        cm_dacr = col_map(hdrs_dacr)
        card_refs = []
        for row in rows_dacr:
            if not any(v is not None for v in row):
                continue
            obj = {k: clean(row[v]) for k, v in cm_dacr.items() if v < len(row) and clean(row[v]) is not None}
            if obj:
                card_refs.append(obj)
        write_json(ref_dir / "daily_astrological_cards_ref.json", {"cards": card_refs})

    # All other reference sheets
    for sheet_name, popup_ref in POPUP_REF_MAP.items():
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠️  Sheet not found: {sheet_name}")
            continue
        ws_ref = wb[sheet_name]
        parsed = parse_reference_sheet(ws_ref, popup_ref_key=popup_ref)
        write_json(ref_dir / f"{sheet_name}.json", parsed)

    # ── Summary ──
    print(f"\n{'─'*50}")
    print(f"✅ Done! Output: {output_dir}/")
    print(f"   📅 {len(all_days)} days → {len(monthly)} monthly + {len(all_days)} day-detail files")
    print(f"   ⭐ {len(ausp_list)} auspicious day entries")
    print(f"   🗓️  {len(ev_list)} events")
    all_files = list(output_dir.rglob("*.json"))
    total_kb = sum(f.stat().st_size for f in all_files) / 1024
    print(f"   📦 {len(all_files)} JSON files, {total_kb:.0f} KB total")
    print(f"{'─'*50}\n")

    # ── Auto-generate manifest ──
    manifest_path = script_dir / "manifest.json"
    try:
        import subprocess
        print("📋 Auto-generating manifest.json ...")
        result = subprocess.run(
            ["python3", str(script_dir / "generate_manifest.py"),
             "--data-dir", str(output_dir),
             "--output",   str(manifest_path)],
            capture_output=True, text=True
        )
        if result.returncode == 0:
            # Print only the summary lines
            for line in result.stdout.splitlines():
                if line.strip():
                    print(f"   {line}")
        else:
            print(f"   ⚠️  manifest generation warning: {result.stderr[:200]}")
    except Exception as e:
        print(f"   ⚠️  Could not auto-generate manifest: {e}")


if __name__ == "__main__":
    main()
